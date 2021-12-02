# -*- coding: utf-8 -*-
"""
Created on Tue Nov 16 18:40:36 2021

@author: Jack
"""

import pandas as pd
import numpy as np
# from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Open an existing workbook:
# wb = load_workbook('D:/DataFileForSeismometer/Seismograph_Data.xlsx')
# wb = load_workbook('./Seismograph_Data.xlsx')

# https://openpyxl.readthedocs.io/en/stable/tutorial.html
# Crate a new notebook
wb = Workbook()
# A workbook is always created with at least one worksheet. You can get it by using the Workbook.active property:
ws = wb.active
# Name worksheet 0
ws.title = "My seimic data"

# exit()

# https://pandas.pydata.org/docs/reference/api/pandas.read_csv.html
# dtypeType name or dict of column -> type, optional
# Data type for data or columns. E.g. {‘a’: np.float64, ‘b’: np.int32, ‘c’: ‘Int64’} Use str or object
# together with suitable na_values settings to preserve and not interpret dtype. If converters are specified,
# they will be applied INSTEAD of dtype conversion.
col_names = ['Date Time', 'Roll', 'Pitch', 'Direction Angle', 'Temperature']
df = pd.read_csv('./DATALOG1.csv', names=col_names, low_memory=False, encoding='utf-8')

print(df)

# https://stackoverflow.com/questions/34962104/how-can-i-use-the-apply-function-for-a-single-column
df['Roll'] = df['Roll'].abs()
df['Pitch'] = df['Pitch'].abs()

# Print elements 1 to 5 only.
print(df.iloc[1:5])

# To convert a dataframe into a worksheet
# https://openpyxl.readthedocs.io/en/stable/pandas.html
for row in dataframe_to_rows(df, index=False, header=True):
    ws.append(row)

# https://openpyxl.readthedocs.io/en/stable/tutorial.html#saving-to-a-file
wb.save('Seismograph_Data.xlsx')
