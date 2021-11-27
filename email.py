import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

wb = load_workbook('D:/DataFileForSeismometer/Seismograph_Data.xlsx')
ws = wb.active

col_names = ['Date Time', 'Roll', 'Pitch', 'Direction Angle', 'Temperature']
df = pd.read_csv('D:/DataFileForSeismometer/DATALOG2.csv', names=col_names, low_memory=False, encoding='utf-8', dtype={"Data Time": str, "Roll": str,
"Pitch": str, "Direction Angle": str, "Temperature": str})

df_1 = pd.DataFrame(df, columns=['Date Time', 'Roll', 'Pitch'])

rows = dataframe_to_rows(df_1)
for r_idx, row in enumerate(rows, 2):
    for c_idx, value in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx, value=value)

# df1['Roll'].abs()

print(df1['Roll'].abs())
